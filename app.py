import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from datetime import datetime
import re

st.set_page_config(page_title="📊 Campaign demand estimation app", layout="wide")

@st.cache_data
def load_excel_and_unmerge(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    for merged_range in list(ws.merged_cells.ranges):
        tl_row = merged_range.min_row
        tl_col = merged_range.min_col
        top_left_value = ws.cell(row=tl_row, column=tl_col).value
        ws.unmerge_cells(range_string=str(merged_range))
        for r in ws.iter_rows(min_row=merged_range.min_row,
                              max_row=merged_range.max_row,
                              min_col=merged_range.min_col,
                              max_col=merged_range.max_col):
            for cell in r:
                cell.value = top_left_value

    for col in ws.columns:
        prev_value = None
        for cell in col:
            if cell.value is not None:
                prev_value = cell.value
            else:
                cell.value = prev_value

    data_iter = ws.values
    try:
        headers = next(data_iter)
    except StopIteration:
        wb.close()
        return pd.DataFrame()
    df = pd.DataFrame(data_iter, columns=headers)
    wb.close()

    df = df.ffill(axis=0)
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace(r'[\u00A0\u202F]', '', regex=True)
    )

    return df


def clean_demand_column(df, demand_col='Demand'):
    def parse_demand(val):
        if pd.isna(val):
            return None
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return float(val)
        s = str(val)
        s = s.replace('\u00A0', '').replace('\u202F', '').replace(' ', '')
        s = re.sub(r'[^\d,.\-]', '', s)
        if s.count(',') == 1 and s.count('.') == 0:
            s = s.replace(',', '.')
        if s == '' or s == '-' or s == '.' or s == ',':
            return None
        try:
            num = float(s)
            if abs(num) > 1e12:
                return None
            return num
        except ValueError:
            return None

    if demand_col in df.columns:
        df[demand_col] = df[demand_col].apply(parse_demand)
        # st.info(f"Demand cleaned — valid: {valid}, invalid: {invalid}")  # <-- linia usunięta
    else:
        st.warning(f"Column '{demand_col}' not found.")
    return df


def filter_data(df, country, search_filter, start_date, end_date, selected_category=None):
    if 'Country' not in df.columns:
        return pd.DataFrame()

    df_filtered = df[df['Country'] == country].copy()

    for col in ['Name', 'Description', 'Category']:
        if col in df_filtered.columns:
            df_filtered[col] = (
                df_filtered[col].astype(str)
                .str.strip()
                .str.replace(r'[\u00A0\u202F]', '', regex=True)
            )

    if selected_category and selected_category != "All" and 'Category' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Category'].str.lower() == selected_category.strip().lower()]

    if search_filter and len(search_filter.strip()) >= 3:
        pattern = search_filter.strip().replace('\u00A0', '').replace('\u202F', '')
        name_mask = df_filtered['Name'].str.contains(pattern, case=False, na=False, regex=False) if 'Name' in df_filtered.columns else pd.Series(False, index=df_filtered.index)
        desc_mask = df_filtered['Description'].str.contains(pattern, case=False, na=False, regex=False) if 'Description' in df_filtered.columns else pd.Series(False, index=df_filtered.index)
        df_filtered = df_filtered[name_mask | desc_mask]

    if 'Start' in df_filtered.columns and 'End' in df_filtered.columns:
        df_filtered['Start'] = pd.to_datetime(df_filtered['Start'], dayfirst=True, errors='coerce')
        df_filtered['End'] = pd.to_datetime(df_filtered['End'], dayfirst=True, errors='coerce')
        start_ts = pd.to_datetime(start_date)
        end_ts = pd.to_datetime(end_date)
        df_filtered = df_filtered[(df_filtered['End'] >= start_ts) & (df_filtered['Start'] <= end_ts)]

    return df_filtered


def estimate_demand(earlier_df, later_df, percentage):
    earlier_mean = earlier_df['Demand'].mean() if (earlier_df is not None and not earlier_df.empty) else 0
    later_mean = later_df['Demand'].mean() if (later_df is not None and not later_df.empty) else 0
    adjusted_earlier = earlier_mean * (1 + percentage / 100)
    if (earlier_df is None or earlier_df.empty) and (later_df is None or later_df.empty):
        return None
    if earlier_df is None or earlier_df.empty:
        return later_mean
    if later_df is None or later_df.empty:
        return adjusted_earlier
    return (adjusted_earlier + later_mean) / 2


def reorder_columns(df):
    cols = df.columns.tolist()
    if 'Name' in cols and 'Description' in cols:
        cols.remove('Description')
        idx = cols.index('Name') + 1
        cols.insert(idx, 'Description')
        return df[cols]
    return df


# ---- Main UI ----
st.title("📊 Campaign demand estimation app")

uploaded_file = st.file_uploader("📂 Upload campaign data Excel file (.xlsx/.xls)", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        raw_bytes = uploaded_file.read()
        df = load_excel_and_unmerge(raw_bytes)

        if df.empty:
            st.error("No data read from Excel.")
        else:
            df.columns = df.columns.astype(str).str.strip().str.replace(r'[\u00A0\u202F]', '', regex=True)
            # linia wyświetlająca kolumny usunięta

            required_cols = {'Country', 'Name', 'Description', 'Start', 'End', 'Demand'}
            missing = required_cols - set(df.columns)
            if missing:
                st.error(f"Missing required columns: {missing}")
            else:
                df = clean_demand_column(df, demand_col='Demand')

                country_list = df['Country'].dropna().unique().tolist()
                selected_country = st.selectbox("🌍 Select country:", country_list)

                categories = df['Category'].dropna().unique().tolist() if 'Category' in df.columns else []
                categories = sorted(categories)
                selected_category = st.selectbox("🏷️ Select category:", ["All"] + categories)

                search_filter = st.text_input("🔎 Search campaigns by name or description (min 3 letters):")

                st.subheader("⏳ Earlier Period")
                earlier_start_date = st.date_input("Start date (Earlier Period):", key='earlier_start')
                earlier_end_date = st.date_input("End date (Earlier Period):", key='earlier_end')

                st.subheader("⏳ Later Period")
                later_start_date = st.date_input("Start date (Later Period):", key='later_start')
                later_end_date = st.date_input("End date (Later Period):", key='later_end')

                st.subheader("📈 Target growth from Earlier Period (%)")
                target_growth = st.number_input(
                    "Enter growth percentage (can be negative):",
                    min_value=-100, max_value=1000, step=1, format="%d"
                )

                earlier_filtered = filter_data(df, selected_country, search_filter, earlier_start_date, earlier_end_date, selected_category)
                later_filtered = filter_data(df, selected_country, search_filter, later_start_date, later_end_date, selected_category)

                earlier_filtered = reorder_columns(earlier_filtered)
                later_filtered = reorder_columns(later_filtered)

                st.subheader("Earlier Period (filtered):")
                st.dataframe(earlier_filtered.head(200))

                st.subheader("Later Period (filtered):")
                st.dataframe(later_filtered.head(200))

                st.subheader("Select campaigns to include from Earlier Period:")
                earlier_selections = {}
                for idx, row in earlier_filtered.iterrows():
                    label = f"{row.get('Name','')} | {row.get('Description','')} | Start: {row.get('Start','')} | End: {row.get('End','')} | Demand: {row.get('Demand','')}"
                    earlier_selections[idx] = st.checkbox(label, value=True, key=f"earlier_{idx}")

                st.subheader("Select campaigns to include from Later Period:")
                later_selections = {}
                for idx, row in later_filtered.iterrows():
                    label = f"{row.get('Name','')} | {row.get('Description','')} | Start: {row.get('Start','')} | End: {row.get('End','')} | Demand: {row.get('Demand','')}"
                    later_selections[idx] = st.checkbox(label, value=True, key=f"later_{idx}")

                earlier_selected_df = earlier_filtered.loc[[i for i,v in earlier_selections.items() if v]] if earlier_selections else pd.DataFrame()
                later_selected_df = later_filtered.loc[[i for i,v in later_selections.items() if v]] if later_selections else pd.DataFrame()

                if st.button("📈 Calculate Estimation"):
                    if earlier_selected_df.empty and later_selected_df.empty:
                        st.warning("No campaigns selected in either period for estimation.")
                    else:
                        estimation = estimate_demand(earlier_selected_df, later_selected_df, target_growth)
                        if estimation is None:
                            st.warning("Unable to calculate estimation with the given data.")
                        else:
                            st.success(f"Estimated Demand: {estimation:,.2f} EUR")
                            st.markdown("### Data used for estimation:")
                            if not earlier_selected_df.empty:
                                st.write("Earlier Period Campaigns:")
                                st.dataframe(earlier_selected_df)
                            if not later_selected_df.empty:
                                st.write("Later Period Campaigns:")
                                st.dataframe(later_selected_df)
                            combined_df = pd.concat([earlier_selected_df, later_selected_df]).drop_duplicates()
                            csv = combined_df.to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label="📥 Download selected campaigns data as CSV",
                                data=csv,
                                file_name='campaign_estimation_data.csv',
                                mime='text/csv'
                            )

    except Exception as e:
        st.error(f"Error processing file: {e}")
